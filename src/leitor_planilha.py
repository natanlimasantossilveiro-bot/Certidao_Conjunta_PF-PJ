from datetime import datetime, date
from openpyxl import load_workbook

COLUNAS_OBRIGATORIAS = ["tipo", "documento", "data_nascimento"]


def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def limpar_documento(valor):
    documento = normalizar_texto(valor)

    if documento.endswith(".0"):
        documento = documento[:-2]

    documento = (
        documento.replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    )

    return documento


def normalizar_data(valor):
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    texto = normalizar_texto(valor)

    if not texto:
        return ""

    formatos_possiveis = [
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]

    for formato in formatos_possiveis:
        try:
            data_convertida = datetime.strptime(texto, formato)
            return data_convertida.strftime("%d/%m/%Y")
        except ValueError:
            pass

    return texto


def validar_colunas(cabecalho):
    colunas_encontradas = [normalizar_texto(coluna).lower() for coluna in cabecalho]

    for coluna in COLUNAS_OBRIGATORIAS:
        if coluna not in colunas_encontradas:
            raise ValueError(f"Coluna obrigatória não encontrada: {coluna}")

    return colunas_encontradas


def ler_planilha_certidoes(caminho_arquivo):
    workbook = load_workbook(caminho_arquivo)
    planilha = workbook.active
    linhas = list(planilha.iter_rows(values_only=True))

    if not linhas:
        raise ValueError("A planilha está vazia.")

    cabecalho = validar_colunas(linhas[0])

    indice_tipo = cabecalho.index("tipo")
    indice_documento = cabecalho.index("documento")
    indice_data_nascimento = cabecalho.index("data_nascimento")

    registros_validos = []
    erros = []

    for numero_linha, linha in enumerate(linhas[1:], start=2):
        tipo = normalizar_texto(linha[indice_tipo]).lower()
        documento = limpar_documento(linha[indice_documento])
        data_nascimento = normalizar_data(linha[indice_data_nascimento])

        if not tipo and not documento and not data_nascimento:
            continue

        erros_linha = []

        if tipo not in ["pf", "pj"]:
            erros_linha.append("tipo deve ser 'pf' ou 'pj'")

        if not documento:
            erros_linha.append("documento é obrigatório")

        if tipo == "pf" and len(documento) != 11:
            erros_linha.append("CPF deve conter 11 dígitos")

        if tipo == "pj" and len(documento) != 14:
            erros_linha.append("CNPJ deve conter 14 dígitos")

        if tipo == "pf" and not data_nascimento:
            erros_linha.append("data_nascimento é obrigatória para PF")

        if erros_linha:
            erros.append({"linha": numero_linha, "erros": erros_linha})
            continue

        registros_validos.append(
            {
                "linha": numero_linha,
                "tipo": tipo,
                "documento": documento,
                "data_nascimento": data_nascimento,
            }
        )

    workbook.close()

    return registros_validos, erros
